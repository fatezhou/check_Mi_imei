import configparser
import requests
from lxml import etree
import os
from urllib import parse
import sys
from ctypes import *
import json
import pdb
from requests.cookies import RequestsCookieJar
from bs4 import BeautifulSoup
import sqlite3
import time
import threading
import math
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
import csv
Msg = []
TaskList = []
global threadCreated
threadCreated=False
threadCreatedStop=False

def GetFileName(fileFormat):
	strdate = (time.strftime('%Y%m%d%H%M%S',time.localtime(time.time())))
	strTime = str(time.time())
	timei = strTime.find(".") +1
	timel = len(strTime)
	strTime = strTime[timei:timel]
	filename = strdate + strTime + "." + fileFormat
	return filename


logfileName = GetFileName("log")
logging.basicConfig(level=logging.INFO,filename=logfileName,filemode='a',format='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s')		
logLines = 0
def WriteLog(logmsg):
	global logLines
	if logLines>5000:
		logfileName = GetFileName("log")
		logging.basicConfig(level=logging.INFO,filename=logfileName,filemode='a',format='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s')
		logLines = 0
	
	logging.info(logmsg)
	logLines = logLines + 1



def ReadCSVFile(fileName,emsiA):
	csv_reader = csv.reader(open(fileName))
	bAdd = False
	for row in csv_reader:		
		TaskList.append(row)	

def WriteCSVFile(name,delay,config):
	#
	lines = 0
	iniPath ='xiaomi_phone.ini'	
	
	filename = GetFileName("csv")
	while 1:
		try:
			
			if lines >= 5000:
				csvFile2.close()	
				lines = 0
				filename = GetFileName("csv")
			
			csvFile2 = open(filename,'a+', newline='') # 设置newline，否则两行之间会空一行
			writer = csv.writer(csvFile2)			
			for li in Msg:				
				lines = lines + 1				
				writer.writerow(li)
			csvFile2.close()	
			Msg.clear()			
			
			config.set("taskrtn","filename",filename)
			config.write(open(iniPath, "w"))
			if threadCreatedStop == True:
				config2 = configparser.ConfigParser()
				config2.read_file(open("stop.ini"))
				config2.set("task","stop","1")
				config2.write(open("stop.ini", "w"))
				break
			time.sleep(int(delay))
		except:
			msg = filename + "文件打开失败,请关闭!"
			print(msg)
			config.set("taskrtn","msg",msg)
			config.write(open(iniPath, "w"))
			csvFile2.close()	
			time.sleep(10)
			WriteCSVFile(name,delay,config)


def WriteExcelFile(name,delay,config):
	lines = 0
	filenumber = 1
	iniPath ='xiaomi_phone.ini'	
	filename = ""
	while 1:
		try:
			filename = "phone"+ str(filenumber) + ".xlsx"
			excel = ""
			try:
				excel = load_workbook(filename)		
			except:
				excel = Workbook()
						
			sheet = excel.get_active_sheet()
			
			for li in Msg:
				sheet.append(li)
				lines = lines + 1
				
			if lines >= 5000:
				lines = 0
				filenumber = filenumber + 1

			#sheet.title = "myPhone"
			excel.save(filename)
			Msg.clear()
			config.set("taskrtn","filename",filename)
			config.write(open(iniPath, "w"))
			if threadCreatedStop == True:
				break
			time.sleep(int(delay))

		except:
			msg = filename + "文件打开失败,请关闭!"
			print(msg)
			config.set("taskrtn","msg",msg)
			config.write(open(iniPath, "w"))

			time.sleep(10)
			WriteExcelFile(name,delay,config)

def GetPhoneMsg(emsi,config):
	try:
		url = "http://www.mi.com/global/verify/#/en/tab/imei"

		session = requests.session()

		resA = session.get(url)
		
		resA.cookies.set("xm_en_sid","9v7735vk36ia391diqk9ihj9h4")
		resA.cookies.set("xmuuid","XMGUEST-B6EF98E8-EA10-DB8A-3F59-151C670A83D4")
		resA.cookies.set("__guid","215952004.2777925070551886000.1540879848019.8833")
		resA.cookies.set("_ga","GA1.2.1975958200.1540897593")
		resA.cookies.set("mstuid","1540897593054_3556")
		resA.cookies.set("monitor_count","36")	
		path = "qrcode.jpg"
		
		strTime = str(time.time()*1000)
		timei = strTime.find(".")		
		strTime = strTime[0:timei]

		host = "https://buy.mi.com/en/other/getimage?_=" + str(strTime)
		QrPicData = session.get(host,cookies=session.cookies,timeout = 60)    
		with open(path,"wb") as f:
			f.write(QrPicData.content)
			f.close()
		
		#getcode
		YDMApi = ""
		try:
			YDMApi = windll.LoadLibrary('yundamaAPI-x64.dll')
		except:
			resData = "云打码出错:dll丢失" + "  停止运行,请获取后，再重新运行！"
			config.set("taskrtn","msg",resData)
			config.write(open(iniPath, "w"))
			#if len(Msg) > 0:
			#	WriteExcelFile("name",1,config)
			threadCreatedStop = True
			return resData

		appId = 6051   
		appKey = b'ed0ef90d97a6b46deb768f320531fa9a'     

		user = config.get("yundama","user")
		pwd = config.get("yundama","pwd")
		username = bytes(user,encoding='utf-8')
		password = bytes(pwd,encoding='utf-8')

		codetype = int(config.get("yundama","codetype"))

		# 分配4个字节存放识别结果
		result = c_char_p(b"    ")    

		# 识别超时时间 单位：秒
		timeout = 60

		# 验证码文件路径
		filename = b'qrcode.jpg'
		
		code=""
		

		# 一键识别函数，无需调用 YDM_SetAppInfo 和 YDM_Login，适合脚本调用
		captchaId = YDMApi.YDM_EasyDecodeByPath(username, password, appId, appKey, filename, codetype, timeout, result)
		if captchaId < 0:
			resData = "云打码出错:" + str(captchaId)
			config.set("taskrtn","msg",resData)
			config.write(open(iniPath, "w"))			
			return resData
		#print("一键识别：验证码ID：%d，识别结果：%s" % (captchaId, result.value))
		code = result.value.decode()
		#print("验证码：" + str(code))
		strTime = str(time.time()*1000)
		timei = strTime.find(".")		
		strTime = strTime[0:timei]
		WriteLog(code)
		host = "http://buy.mi.com/en/other/checkimei?jsonpcallback=JSON_CALLBACK&keyword="+str(emsi)+"&vcode="+code+"&v=0.47245648708541066&_="+str(strTime)
		WriteLog(host)
		reqA = session.get(host,cookies = session.cookies)				   
		res = reqA.text
		#res = 'if(window.JSON_CALLBACK)JSON_CALLBACK({"code":1,"msg":"suc","data":{"goods_name":"Redmi 4X 3GB RAM 32GB ROM Black","add_time":1511478067,"country_text":"Bangladesh"}});'
		resLen = len(res) - 2
		index = res.find("{")
		jsonRes = res[index:resLen]
		
		return jsonRes
	except:
		GetPhoneMsg(emsi,config)


def ExecPiLiang(data,config):
	try:
		data = data.replace(' ','')
		data = data.replace('\t','').strip()	
		msgRes = GetPhoneMsg(data,config)
		print(data)	
		print(msgRes)				
		if msgRes.find("停止运行") != -1:
			pass			
		else:		
			resJson = json.loads(msgRes)
			rescode = resJson['code']	
			
			if rescode == 1:							
				resres = json.dumps(resJson['data'])
				dataList = []
				dataList.append(data)
				dataList.append(resJson['data']['goods_name'])
				dataList.append(resJson['data']['country_text'])
				dataList.append(resJson['data']['add_time'])				
				Msg.append(dataList)
			elif rescode == 70011:
				ExecPiLiang(data,config)
	
	except:
		ExecPiLiang(data,config)



def main(emsiA):	
	try:				
		global threadCreatedStop
		curPath = os.getcwd()
		iniPath ='xiaomi_phone.ini'		
		config = configparser.ConfigParser()
		config.read_file(open(iniPath))								

		global threadCreated
		if threadCreated==False:			
			getPriceThread = threading.Thread(target=WriteCSVFile, args=("WriteCSVFile",30,config,) )
			getPriceThread.start()
			threadCreated=True		

		
		taskType = config.get("task","tasktype")
		times = 1
		if taskType == "1":			
			if emsiA == "":
				emsi = config.get("task","emsi")	
			else:
				emsi = emsiA

			if emsi == "":
				config.set("taskrtn","msg","停止运行,请填写emsi或者s/n码后重新运行!")
				config.write(open(iniPath, "w"))
				threadCreatedStop = True
			else :
				while 1:	
					config2 = configparser.ConfigParser()
					config2.read_file(open("stop.ini"))				
					bstop = config2.get("task","stop")
					print(bstop)
					if bstop == "1":	
						config2.set("taskrtn","emsi",emsi)
						config2.write(open("stop.ini", "w"))
						#if len(Msg) > 0:
						#	WriteExcelFile("name",1,config)
						threadCreatedStop = True		
						return
					emsiA = emsi
					msgRes = GetPhoneMsg(emsi,config)	
					if msgRes == None:
						continue
					print(emsi)	
					print(msgRes)	
					WriteLog(msgRes)		
					if msgRes.find("停止运行") != -1:
						config.set("taskrtn","emsi",emsi)
						config.write(open(iniPath, "w"))
						pass
					else:
						resJson = json.loads(msgRes)
						rescode = resJson['code']	
						
						if rescode == 1:							
							resres = json.dumps(resJson['data'])
							dataList = []
							dataList.append(str(emsi))
							dataList.append(resJson['data']['goods_name'])
							dataList.append(resJson['data']['country_text'])
							dataList.append(resJson['data']['add_time'])
							
							Msg.append(dataList)							
							emsi = int(emsi) + 7		
							times = 1					
						elif rescode == 70011:
							pass
						else:
							if times == 1:
								emsi = int(emsi) + 1
								times = times + 1
								WriteLog("+1")
							else:
								emsi = int(emsi) + 9
								times = 1
								WriteLog("+9")
												

		else:
			fileName = config.get("task","filename")
				
			ReadCSVFile(fileName,emsiA)
			print(fileName)						
			if True:
				for row in TaskList:
					for l in row:							
						ExecPiLiang(l,config)
								

			else:
				try:
					excel = load_workbook(fileName)		
				except:
					config.set("taskrtn","msg","导入的emsi文件不存在")
					config.write(open(iniPath, "w"))
					#if len(Msg) > 0:
						#WriteExcelFile("name",1,config)
					threadCreatedStop = True
					return
				sheet = excel.get_active_sheet()
				columnSize = sheet.max_column
				rowSize = sheet.max_row
				i = 1
				while i <= columnSize:
					j = 1				
					while j<= rowSize:
						cellvalue = sheet.cell(row=j, column=i)
						if cellvalue.value != None:			
							emsiA = cellvalue.value			
							msgRes = GetPhoneMsg(cellvalue.value,config)							
							print(cellvalue.value)	
							print(msgRes)				
							if msgRes.find("停止运行") != -1:
								pass
								
							else:
								resJson = json.loads(msgRes)
								rescode = resJson['code']	
								
								if rescode == 1:							
									resres = json.dumps(resJson['data'])
									dataList = []
									dataList.append(cellvalue.value)
									dataList.append(resJson['data']['goods_name'])
									dataList.append(resJson['data']['country_text'])
									dataList.append(resJson['data']['add_time'])
									
									Msg.append(dataList)							
																														
						j = j+ 1
					i = i + 1	
			#if len(Msg) > 0:
				#WriteExcelFile("name",1,config)
			threadCreatedStop = True						

		
			
	except:		
		print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))
		print("执行中断，重新执行！") 	
		WriteLog("执行中断，重新执行！")
		main(emsiA)			
		

main("")




